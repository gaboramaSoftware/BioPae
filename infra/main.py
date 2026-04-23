import os
import asyncio
import hashlib
import logging
import random
import shutil
from datetime import datetime
from contextlib import asynccontextmanager
from fastapi import FastAPI, HTTPException, BackgroundTasks, WebSocket, WebSocketDisconnect, UploadFile, File, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
from pydantic import BaseModel
from typing import Optional
from infra.DB import modelos
import uvicorn
from sqlalchemy import text  # NUEVO: Importación necesaria para el Punto 5

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

from core.Services.userServces.Atributos.HuellaService import HuellaService
from infra.Controller.Alumno.huellaController import HuellaController
from infra.Controller.RegistrosController import RegistrosController
from core.Services.ticketServices.horarioService import obtener_tipo_racion, descripcion_horario
from core.Domain.Repository.UserRepository import SessionLocal, crearUsuarioBase, eliminarUsuario
from infra.DB.modelos import Usuario, Curso, AdminConfig

def _crear_backup_db():
    """Crea un backup con timestamp de la base de datos. Mantiene los últimos 48 backups."""
    from core.Domain.Repository.IniciarDB import buscarDB
    ruta_db = buscarDB()
    if not ruta_db or not os.path.exists(ruta_db):
        logger.warning("[BACKUP] No se encontró la base de datos para respaldar.")
        return

    carpeta_backup = os.path.join(os.path.dirname(ruta_db), "backups")
    os.makedirs(carpeta_backup, exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    nombre_backup = f"biopae_{timestamp}.db"
    ruta_destino = os.path.join(carpeta_backup, nombre_backup)

    shutil.copy2(ruta_db, ruta_destino)
    logger.info(f"[BACKUP] Backup creado: {ruta_destino}")

    # Mantener solo los últimos 48 backups (2 días a 1 backup/hora)
    backups = sorted([
        os.path.join(carpeta_backup, f)
        for f in os.listdir(carpeta_backup)
        if f.startswith("biopae_") and f.endswith(".db")
    ])
    for viejo in backups[:-48]:
        os.remove(viejo)
        logger.info(f"[BACKUP] Backup antiguo eliminado: {viejo}")

async def _tarea_backup_periodico():
    """Crea un backup cada hora mientras el servidor esté corriendo."""
    while True:
        await asyncio.sleep(3600)
        _crear_backup_db()

@asynccontextmanager
async def lifespan(app: FastAPI):
    """Al arrancar el servidor, inyecta las huellas de SQLite en la RAM del sensor."""
    logger.info("[STARTUP] Iniciando servidor. Cargando huellas persistidas en RAM del sensor...")
    exito, msg = hardware_service.inicializar()
    if exito:
        resultado = hardware_service.cargar_huellas_iniciales()
        logger.info(f"[STARTUP] Inyección finalizada: {resultado}")
    else:
        logger.warning(f"[STARTUP] Sensor no disponible al arrancar ({msg}). Las huellas no se cargaron.")

    tarea_backup = asyncio.create_task(_tarea_backup_periodico())

    yield

    # Shutdown: backup final antes de cerrar
    tarea_backup.cancel()
    _crear_backup_db()
    logger.info("[SHUTDOWN] Backup de cierre creado correctamente.")

app = FastAPI(lifespan=lifespan)

class _NoCacheHTMLMiddleware(BaseHTTPMiddleware):
    """Evita que el navegador cachee archivos .html y .js del frontend."""
    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        path = request.url.path
        if path.endswith('.html') or path.endswith('.js'):
            response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
        return response

app.add_middleware(_NoCacheHTMLMiddleware)

def _sort_key_curso_chileno(usuario):
    """Clave de ordenación por sistema educacional chileno:
    Pre-Kinder(0) → Kinder(1) → 1°..8° Básico(2-9) → 1°..4° Medio(10-13),
    luego letra A-Z, luego nombre A-Z."""
    c = usuario.curso
    if not c:
        return (9999, '', (usuario.nombre or '').lower())
    nivel  = c.nivel  or ''
    numero = c.numero or 0
    letra  = (c.letra or '').upper()
    nombre = (usuario.nombre or '').lower()
    if   nivel == 'Pre-Kinder': nivel_num = 0
    elif nivel == 'Kinder':     nivel_num = 1
    elif nivel == 'Basico':     nivel_num = 1 + numero   # 2-9
    elif nivel == 'Medio':      nivel_num = 9 + numero   # 10-13
    else:                       nivel_num = 9999
    return (nivel_num, letra, nombre)

def _seed_cursos():
    """Puebla la tabla cursos con todos los niveles A-F si está vacía."""
    db = SessionLocal()
    try:
        if db.query(Curso).count() > 0:
            return
        cursos_base = (
            [(None, "Pre-Kinder"), (None, "Kinder")] +
            [(i, "Basico") for i in range(1, 9)] +
            [(i, "Medio")  for i in range(1, 5)]
        )
        letras = list("A")
        for numero, nivel in cursos_base:
            for letra in letras:
                db.add(Curso(numero=numero, nivel=nivel, letra=letra))
        db.commit()
    finally:
        db.close()

_seed_cursos()

def _hash_password(password: str) -> str:
    return hashlib.sha256(password.encode('utf-8')).hexdigest()

def _seed_admin_config():
    """Crea las credenciales de administrador por defecto si no existen."""
    db = SessionLocal()
    try:
        if db.query(AdminConfig).count() > 0:
            return
        db.add(AdminConfig(rut="11111111", password_hash=_hash_password("admin123")))
        db.commit()
        logger.info("[SEED] Credenciales admin por defecto creadas.")
    finally:
        db.close()

_seed_admin_config()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ================================
# CONFIGURACIÓN DE HARDWARE Y CONTROLADORES
# ================================

# FIX PUNTO 6: Leer el TOTEM_ID desde las variables de entorno (por defecto 1)
TOTEM_ID = int(os.getenv("TOTEM_ID", "1"))

hardware_service = HuellaService()

# FIX PUNTO 1 (Continuación): Inyectamos la conexión a la base de datos
huella_buffer = HuellaController(db=SessionLocal())



# Se inicializa con el TOTEM_ID dinámico
registros_controller = RegistrosController(totem_id=TOTEM_ID)

static_dir = os.path.join(os.path.dirname(__file__), "..", "Frontend", "src")
os.makedirs(static_dir, exist_ok=True)

@app.get("/")
def serve_frontend():
    return FileResponse(os.path.join(static_dir, "index.html"))

# ================================
# MODELOS PYDANTIC
# ================================
class NuevoUsuario(BaseModel):
    nombre: str
    rut: Optional[str] = None
    curso_id: Optional[int] = None
    curso_nombre: Optional[str] = None
    estado_id: int
    es_pae: bool = False

class EdicionHuella(BaseModel):
    id_usuario: int

class ProcesarTicket(BaseModel):
    usuario_id: int

class UpdateUser(BaseModel):
    nombre: Optional[str] = None
    rut: Optional[str] = None
    es_pae: Optional[bool] = None
    curso_id: Optional[int] = None
    observaciones: Optional[str] = None

class VincularHuella(BaseModel):
    usuario_id: int
    huella_hex: str

# ================================
# ENDPOINTS ESTADO DEL SISTEMA
# ================================

@app.get("/api/db/status")
def db_status():
    db = SessionLocal()
    try:
        # FIX PUNTO 5: Envolver el query en text() para compatibilidad con SQLAlchemy 2.0
        db.execute(text("SELECT 1"))
        return {"estado": True, "mensaje": "Base de datos conectada"}
    except Exception as e:
        return {"estado": False, "mensaje": str(e)}
    finally:
        db.close()

@app.get("/api/sensor/status")
def api_sensor_status():
    exito, msg = hardware_service.inicializar()
    if exito and not hardware_service._huellas_cargadas:
        resultado = hardware_service.cargar_huellas_iniciales()
        logger.info(f"[SENSOR] Huellas recargadas tras reconexión: {resultado}")
    return {"available": exito, "status": "ready" if exito else "error", "detalle": msg}

@app.get("/api/sensor/verificar")
def api_sensor_verificar():
    """Verifica si el hardware físico del sensor está conectado (close + init real).
    Más confiable que /status para detectar desconexión del dispositivo USB."""
    conectado = hardware_service.verificar_conexion_hardware()
    return {"available": conectado}

@app.get("/api/horario/status")
def api_horario_status():
    tipo = obtener_tipo_racion()
    return {
        "en_horario": tipo is not None,
        "tipo_racion": tipo,
        "descripcion": descripcion_horario()
    }

# ================================
# ENDPOINTS BIOMETRÍA
# ================================

@app.post("/api/huella/enrolar")
def api_enrolar(datos: NuevoUsuario, background_tasks: BackgroundTasks):
    usuario = crearUsuarioBase(datos.nombre, datos.rut, datos.curso_id, datos.estado_id, datos.es_pae)
    if not usuario:
        raise HTTPException(status_code=500, detail="Error al crear el usuario en la base de datos")

    exito_init, msg = hardware_service.inicializar()
    if not exito_init:
        eliminarUsuario(usuario.id)
        raise HTTPException(status_code=500, detail=msg)

    background_tasks.add_task(
        huella_buffer.procesar_contexto,
        contexto="enrolar",
        controlador_hardware=hardware_service,
        id_alumno=usuario.id
    )
    return {"estado": True, "mensaje": "Esperando huella en el sensor...", "id_usuario": usuario.id}

@app.put("/api/huella/editar")
def api_editar_huella(datos: EdicionHuella, background_tasks: BackgroundTasks):
    exito_init, msg = hardware_service.inicializar()
    if not exito_init:
        raise HTTPException(status_code=500, detail=msg)

    background_tasks.add_task(
        huella_buffer.procesar_contexto,
        contexto="editar",
        controlador_hardware=hardware_service,
        id_alumno=datos.id_usuario
    )
    return {"estado": True, "mensaje": "Sensor activado, esperando nueva huella..."}

@app.post("/api/totem/acceso")
def api_totem_acceso(background_tasks: BackgroundTasks):
    """Activa el sensor. El tipo de ración se determina automáticamente por horario."""
    tipo = obtener_tipo_racion()
    if tipo is None:
        raise HTTPException(status_code=403, detail=descripcion_horario())

    exito_init, msg = hardware_service.inicializar()
    if not exito_init:
        raise HTTPException(status_code=500, detail=msg)

    huella_buffer.limpiar()
    background_tasks.add_task(
        huella_buffer.procesar_contexto,
        contexto="ticket",
        controlador_hardware=hardware_service
    )
    return {"estado": True, "mensaje": f"Sensor activado. Horario: {tipo}", "tipo_racion": tipo}

@app.post("/api/totem/verificar-usuario")
def api_verificar_usuario(background_tasks: BackgroundTasks):
    """Activa el sensor para verificar identidad sin restricción de horario. Usado en flujo de registro."""
    exito_init, msg = hardware_service.inicializar()
    if not exito_init:
        raise HTTPException(status_code=500, detail=msg)

    huella_buffer.limpiar()
    background_tasks.add_task(
        huella_buffer.procesar_contexto,
        contexto="ticket",
        controlador_hardware=hardware_service
    )
    return {"estado": True, "mensaje": "Sensor activado para verificación"}

@app.get("/api/huella/pooling")
def api_consultar_buffer():
    """El frontend llama cada 1 segundo para saber si el sensor terminó."""
    datos = huella_buffer.obtener_datos()
    if datos["hay_datos"]:
        respuesta = dict(datos)
        huella_buffer.limpiar()
        return {"estado": True, "datos": respuesta}
    return {"estado": False, "mensaje": "Aún esperando hardware"}

@app.post("/api/ticket/procesar")
def api_procesar_ticket(datos: ProcesarTicket):
    """
    Llamado por el frontend tras recibir user_id del pooling.
    Valida PAE, horario, duplicados y genera el ticket en BD.
    """
    resultado = registros_controller.procesarAsistencia(datos.usuario_id)
    return resultado

# ================================
# ENDPOINTS MANTENEDOR USUARIOS
# ================================

@app.get("/api/usuarios/buscar")
def buscar_usuario(nombre: str, curso_id: Optional[int] = None):
    """
    Busca alumnos PAE pre-cargados por nombre (parcial) y opcionalmente curso.
    Usado en el flujo de auto-registro cuando la huella no es reconocida.
    """
    db = SessionLocal()
    try:
        query = db.query(Usuario).filter(
            Usuario.nombre.ilike(f"%{nombre}%")
        )
        if curso_id is not None:
            query = query.filter(Usuario.curso_id == curso_id)
        usuarios = query.all()
        return [
            {
                "id": u.id,
                "nombre": u.nombre,
                "rut": u.rut,
                "curso": (" ".join(p for p in [str(u.curso.numero) if u.curso.numero is not None else "", u.curso.nivel, u.curso.letra or ""] if p).strip()) if u.curso else "Sin curso",
                "tiene_huella": u.huella is not None,
            }
            for u in usuarios
        ]
    finally:
        db.close()

@app.get("/api/cursos")
def obtener_cursos():
    db = SessionLocal()
    try:
        cursos = db.query(Curso).all()

        def nombre_curso(c):
            partes = [str(c.numero) if c.numero is not None else "", c.nivel, c.letra or ""]
            return " ".join(p for p in partes if p).strip()
        return [{"id": c.id, "numero": c.numero, "nivel": c.nivel, "letra": c.letra, "nombre": nombre_curso(c)} for c in cursos]
    finally:
        db.close()

@app.delete("/api/registros")
def eliminar_todos_registros():
    from infra.DB.modelos import Registro
    db = SessionLocal()
    try:
        db.query(Registro).delete(synchronize_session=False)
        db.commit()
        return {"estado": True, "mensaje": "Todos los registros han sido eliminados"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()

@app.get("/api/registros")
def obtener_registros(fecha: str = None):
    from datetime import date
    from infra.DB.modelos import Ticket, Registro
    db = SessionLocal()
    try:
        q = db.query(Registro).order_by(Registro.fecha.desc(), Registro.hora.desc())
        if fecha:
            q = q.filter(Registro.fecha == fecha)
        registros = q.limit(1000).all()
        resultado = []
        for r in registros:
            usuario = r.usuario
            curso_nombre = "Sin curso"
            if usuario and usuario.curso:
                c = usuario.curso
                partes = [str(c.numero) if c.numero else "", c.nivel, c.letra or ""]
                curso_nombre = " ".join(p for p in partes if p).strip()
            resultado.append({
                "id": r.id,
                "estudiante": usuario.nombre if usuario else "Desconocido",
                "curso": curso_nombre,
                "racion": r.tipo_solicitud,
                "hora": str(r.hora) if r.hora else "",
                "fecha": str(r.fecha) if r.fecha else "",
                "terminal": r.totem.ubicacion if r.totem else str(r.totem_id),
                "estado": r.estado_registro,
            })
        return resultado
    finally:
        db.close()

@app.get("/api/registros/hoy")
def registros_hoy():
    from datetime import date
    from infra.DB.modelos import Ticket, Registro
    db = SessionLocal()
    hoy = date.today()
    try:
        desayunos = db.query(Ticket).filter(Ticket.fecha_emision == hoy, Ticket.tipo_ticket == "desayuno").count()
        almuerzos = db.query(Ticket).filter(Ticket.fecha_emision == hoy, Ticket.tipo_ticket == "almuerzo").count()
        return {"desayunos": desayunos, "almuerzos": almuerzos, "total": desayunos + almuerzos}
    finally:
        db.close()

@app.delete("/api/usuarios/{user_id}")
def eliminar_usuario_endpoint(user_id: int):
    from core.Domain.Repository.UserRepository import eliminarUsuario
    exito = eliminarUsuario(user_id)
    if not exito:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return {"estado": True, "mensaje": "Usuario eliminado"}

@app.delete("/api/usuarios")
def eliminar_todos_usuarios():
    from infra.DB.modelos import Huella
    db = SessionLocal()
    try:
        db.query(Huella).delete(synchronize_session=False)
        db.query(Usuario).delete(synchronize_session=False)
        db.commit()
        return {"estado": True, "mensaje": "Todos los alumnos y sus huellas han sido eliminados"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()

@app.get("/api/usuarios/{user_id}/historial")
def historial_usuario(user_id: int):
    from infra.DB.modelos import Ticket
    db = SessionLocal()
    try:
        tickets = db.query(Ticket).filter(Ticket.usuario_id == user_id).order_by(Ticket.fecha_emision.desc()).all()
        return [{"fecha": str(t.fecha_emision), "hora": str(t.hora_emision), "tipo": t.tipo_ticket} for t in tickets]
    finally:
        db.close()

@app.get("/api/usuarios")
def obtener_usuarios():
    db = SessionLocal()
    try:
        usuarios = db.query(Usuario).all()
        lista = []
        for u in usuarios:
            curso_nombre = (" ".join(p for p in [str(u.curso.numero) if u.curso.numero is not None else "", u.curso.nivel, u.curso.letra or ""] if p).strip()) if u.curso else "Sin curso"
            lista.append({
                "id": u.id,
                "nombre": u.nombre,
                "rut": u.rut,
                "curso": curso_nombre,
                "curso_id": u.curso_id,
                "es_pae": u.es_pae,
                "tiene_huella": u.huella is not None,
                "observaciones": u.observaciones or ""
            })
        return lista
    finally:
        db.close()

@app.put("/api/usuarios/{user_id}")
def editar_usuario(user_id: int, datos: UpdateUser):
    db = SessionLocal()
    try:
        usuario = db.query(Usuario).filter(Usuario.id == user_id).first()
        if not usuario:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")
        if datos.nombre is not None:
            usuario.nombre = datos.nombre
        if datos.rut is not None:
            usuario.rut = datos.rut
        if datos.es_pae is not None:
            usuario.es_pae = datos.es_pae
        if datos.curso_id is not None:
            usuario.curso_id = datos.curso_id
        if datos.observaciones is not None:
            usuario.observaciones = datos.observaciones[:1000]
        db.commit()
        return {"estado": True, "mensaje": "Usuario actualizado"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()

# ================================
# ENDPOINTS RACIONES JUNAEB
# ================================

@app.get("/api/raciones")
def obtener_raciones():
    from infra.DB.modelos import RacionesConfig, Ticket
    db = SessionLocal()
    try:
        resultado = []
        for tipo in ("desayuno", "almuerzo"):
            config = db.query(RacionesConfig).filter(RacionesConfig.tipo == tipo).first()
            total = config.total if config else 0
            from datetime import date as _date
            usadas = db.query(Ticket).filter(
                Ticket.tipo_ticket == tipo,
                Ticket.fecha_emision == _date.today()
            ).count()
            resultado.append({"tipo": tipo, "total": total, "usadas": usadas, "disponibles": max(0, total - usadas) if total > 0 else None})
        return resultado
    finally:
        db.close()

class ActualizarRaciones(BaseModel):
    total: int

@app.put("/api/raciones/{tipo}")
def actualizar_raciones(tipo: str, datos: ActualizarRaciones):
    from infra.DB.modelos import RacionesConfig
    if tipo not in ("desayuno", "almuerzo"):
        raise HTTPException(status_code=400, detail="Tipo debe ser 'desayuno' o 'almuerzo'")
    if datos.total < 0:
        raise HTTPException(status_code=400, detail="El total no puede ser negativo")
    db = SessionLocal()
    try:
        config = db.query(RacionesConfig).filter(RacionesConfig.tipo == tipo).first()
        if not config:
            from infra.DB.modelos import RacionesConfig as RC
            db.add(RC(tipo=tipo, total=datos.total))
        else:
            config.total = datos.total
        db.commit()
        return {"estado": True, "mensaje": f"Raciones de {tipo} actualizadas a {datos.total}"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()

# ================================
# ENDPOINTS WEBSOCKET (PUSH MODEL)
# ================================

def _resolver_curso_id(datos: NuevoUsuario) -> int:
    """Resuelve el curso_id desde curso_id directo o desde curso_nombre.
    Si no existe en la DB, lo crea. Ejemplo: '1 Basico A' → nivel='1 Basico', letra='A'
    """
    if datos.curso_id is not None:
        return datos.curso_id

    if not datos.curso_nombre:
        raise HTTPException(status_code=400, detail="Se requiere curso_id o curso_nombre")

    # Formato esperado: "1 Basico A" o "Kinder A"
    partes = datos.curso_nombre.strip().split()
    if len(partes) < 2:
        raise HTTPException(status_code=400, detail=f"Formato de curso inválido: '{datos.curso_nombre}'")

    letra_txt = partes[-1].upper()
    resto = partes[:-1]  # ej: ["1", "Basico"] o ["Kinder"]

    numero_txt = None
    if resto[0].isdigit():
        numero_txt = int(resto[0])
        nivel_txt = " ".join(resto[1:])
    else:
        nivel_txt = " ".join(resto)

    db = SessionLocal()
    try:
        curso = db.query(Curso).filter(
            Curso.numero == numero_txt, Curso.nivel == nivel_txt, Curso.letra == letra_txt
        ).first()
        if curso:
            return curso.id
        nuevo = Curso(numero=numero_txt, nivel=nivel_txt, letra=letra_txt)
        db.add(nuevo)
        db.commit()
        db.refresh(nuevo)
        return nuevo.id
    finally:
        db.close()

def _generar_rut_provisional() -> str:
    """Genera un RUT chileno válido de forma aleatoria para uso provisional."""
    base = random.randint(10_000_000, 99_999_999)
    suma, multiplicador, n = 0, 2, base
    while n > 0:
        suma += (n % 10) * multiplicador
        n //= 10
        multiplicador = 2 if multiplicador == 7 else multiplicador + 1
    dv_calc = 11 - (suma % 11)
    dv = '0' if dv_calc == 11 else 'K' if dv_calc == 10 else str(dv_calc)
    return f"{base}-{dv}"

@app.post("/api/usuarios/base")
def crear_usuario_base_endpoint(datos: NuevoUsuario):
    """Crea el usuario en BD sin iniciar el hardware. Usar antes del WS de enrolamiento.
    Si ya existe un usuario con el mismo nombre+curso, retorna el existente con duplicado=True."""
    curso_id = _resolver_curso_id(datos)

    db = SessionLocal()
    try:
        existente_por_nombre = db.query(Usuario).filter(
            Usuario.nombre.ilike(datos.nombre.strip()),
            Usuario.curso_id == curso_id
        ).first()

        existente_por_rut = None
        if datos.rut:
            existente_por_rut = db.query(Usuario).filter(Usuario.rut == datos.rut).first()

        # RUT pertenece a un estudiante distinto → bloquear
        if (existente_por_rut and existente_por_nombre
                and existente_por_rut.id != existente_por_nombre.id):
            raise HTTPException(
                status_code=409,
                detail=f"El RUT {datos.rut} ya pertenece a otro estudiante: {existente_por_rut.nombre}"
            )
        if existente_por_rut and not existente_por_nombre:
            # Mismo estudiante pero en distinto curso registrado → verificar que no sea otro
            existente_por_nombre = existente_por_rut

        existente = existente_por_nombre
        if existente:
            # Sobrescribir datos con los del nuevo registro
            existente.curso_id = curso_id
            existente.estado_id = datos.estado_id
            if datos.rut:
                existente.rut = datos.rut
            if datos.es_pae:
                existente.es_pae = True
            db.commit()
            return {
                "id": existente.id,
                "nombre": existente.nombre,
                "rut": existente.rut,
                "duplicado": True,
                "tiene_huella": existente.huella is not None
            }
    finally:
        db.close()

    rut_final = datos.rut if datos.rut else _generar_rut_provisional()
    usuario = crearUsuarioBase(datos.nombre, rut_final, curso_id, datos.estado_id, datos.es_pae)
    if not usuario:
        raise HTTPException(status_code=500, detail="Error al crear el usuario en la base de datos")
    return {"id": usuario.id, "nombre": usuario.nombre, "duplicado": False}

@app.websocket("/ws/totem")
async def ws_totem(websocket: WebSocket):
    """Flujo completo: identifica huella y procesa ticket. Reemplaza POST /api/totem/acceso + polling."""
    await websocket.accept()
    try:
        tipo = obtener_tipo_racion()
        if tipo is None:
            await websocket.send_json({"estado": "Rechazado", "mensaje": descripcion_horario()})
            return

        if not hardware_service.esta_listo:
            exito, msg = hardware_service.inicializar()
            if not exito:
                await websocket.send_json({"estado": "error", "mensaje": msg})
                return

        huella_buffer.limpiar()

        loop = asyncio.get_running_loop()
        try:
            user_id, score = await asyncio.wait_for(
                loop.run_in_executor(None, hardware_service.identificar_usuario),
                timeout=20.0
            )
        except asyncio.TimeoutError:
            await websocket.send_json({"estado": "Rechazado", "mensaje": "Tiempo de espera agotado"})
            return

        if user_id < 0:
            # Captura fallida (ruido/contacto parcial): cerrar sin mensaje.
            # El frontend detecta el cierre y reintenta automáticamente.
            return

        if user_id == 0:
            # Huella capturada correctamente pero no pertenece a ningún alumno enrolado.
            await websocket.send_json({
                "estado": "Rechazado",
                "mensaje": "Alumno no enrolado en el sistema"
            })
            return

        resultado = registros_controller.procesarAsistencia(user_id)
        await websocket.send_json(resultado)

    except WebSocketDisconnect:
        pass

@app.websocket("/ws/huella/capturar-identificar")
async def ws_huella_capturar_identificar(websocket: WebSocket):
    """Captura la huella, intenta identificar y retorna también los bytes. Usado en flujo de registro."""
    await websocket.accept()
    try:
        if not hardware_service.esta_listo:
            await websocket.send_json({"estado": False, "mensaje": "Sensor no inicializado"})
            return

        huella_buffer.limpiar()

        loop = asyncio.get_running_loop()
        try:
            user_id, score, huella_bytes = await asyncio.wait_for(
                loop.run_in_executor(None, hardware_service.capturar_y_identificar),
                timeout=10.0
            )
        except asyncio.TimeoutError:
            await websocket.send_json({"estado": False, "mensaje": "Tiempo de espera agotado"})
            return

        huella_hex = huella_bytes.hex() if huella_bytes else None
        await websocket.send_json({
            "estado": True,
            "user_id": user_id,
            "score": score,
            "huella_hex": huella_hex
        })

    except WebSocketDisconnect:
        pass

@app.websocket("/ws/huella/identificar")
async def ws_huella_identificar(websocket: WebSocket):
    """Identifica al usuario por huella sin procesar ticket. Usado en flujo de registro del tótem."""
    await websocket.accept()
    try:
        if not hardware_service.esta_listo:
            await websocket.send_json({"estado": False, "mensaje": "Sensor no inicializado"})
            return

        huella_buffer.limpiar()

        loop = asyncio.get_running_loop()
        try:
            user_id, score = await asyncio.wait_for(
                loop.run_in_executor(None, hardware_service.identificar_usuario),
                timeout=10.0
            )
        except asyncio.TimeoutError:
            await websocket.send_json({"estado": False, "mensaje": "Tiempo de espera agotado"})
            return

        await websocket.send_json({"estado": True, "user_id": user_id, "score": score})

    except WebSocketDisconnect:
        pass

@app.websocket("/ws/huella/enrolar/{user_id}")
async def ws_huella_enrolar(websocket: WebSocket, user_id: int):
    """Captura y guarda huella para un usuario ya creado en BD."""
    await websocket.accept()
    try:
        if not hardware_service.esta_listo:
            await websocket.send_json({"estado": False, "mensaje": "Sensor no inicializado"})
            return

        loop = asyncio.get_running_loop()
        try:
            resultado = await asyncio.wait_for(
                loop.run_in_executor(
                    None,
                    lambda: huella_buffer.procesar_contexto(
                        contexto="enrolar",
                        controlador_hardware=hardware_service,
                        id_alumno=user_id
                    )
                ),
                timeout=15.0
            )
        except asyncio.TimeoutError:
            await websocket.send_json({"estado": False, "mensaje": "Tiempo de espera agotado"})
            return

        await websocket.send_json(resultado)

    except WebSocketDisconnect:
        pass

@app.websocket("/ws/huella/editar/{user_id}")
async def ws_huella_editar(websocket: WebSocket, user_id: int):
    """Actualiza la huella de un usuario existente."""
    await websocket.accept()
    try:
        if not hardware_service.esta_listo:
            await websocket.send_json({"estado": False, "mensaje": "Sensor no inicializado"})
            return

        loop = asyncio.get_running_loop()
        try:
            resultado = await asyncio.wait_for(
                loop.run_in_executor(
                    None,
                    lambda: huella_buffer.procesar_contexto(
                        contexto="editar",
                        controlador_hardware=hardware_service,
                        id_alumno=user_id
                    )
                ),
                timeout=15.0
            )
        except asyncio.TimeoutError:
            await websocket.send_json({"estado": False, "mensaje": "Tiempo de espera agotado"})
            return

        await websocket.send_json(resultado)

    except WebSocketDisconnect:
        pass

#endpoints de exportacion

@app.get("/api/exportar/alumnos/filtrado")
def exportar_alumnos_filtrado(cursos: Optional[str] = None):
    """Genera y descarga un xlsx de alumnos filtrado por cursos (string separado por |)."""
    import openpyxl
    from fastapi.responses import StreamingResponse
    import io

    lista_cursos = [c.strip() for c in cursos.split('|')] if cursos else []

    db = SessionLocal()
    try:
        usuarios = sorted(db.query(Usuario).all(), key=_sort_key_curso_chileno)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Alumnos"
        ws.append(["RUT", "Nombre", "Curso", "Estado", "PAE"])

        for u in usuarios:
            c = u.curso
            curso_nombre = (" ".join(p for p in [str(c.numero) if c and c.numero else "", c.nivel if c else "", c.letra or "" if c else ""] if p).strip()) if c else "Sin curso"
            if lista_cursos and curso_nombre not in lista_cursos:
                continue
            estado = u.estado.nombre_estado if u.estado else "Sin estado"
            ws.append([u.rut or "", u.nombre, curso_nombre, estado, "Sí" if u.es_pae else "No"])

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        from datetime import date
        nombre_archivo = f"alumnos_{date.today()}.xlsx"
        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"}
        )
    finally:
        db.close()

@app.post("/api/exportar/alumnos")
async def api_exportar_alumnos():
    loop = asyncio.get_running_loop()
    resultado = await loop.run_in_executor(None, registros_controller.exportar_alumnos_excel)
    if resultado["estado"]:
        return {"estado": True, "mensaje": "Excel generado exitosamente", "ruta": resultado["ruta"]}
    else:
        raise HTTPException(status_code=500, detail=resultado["mensaje"])

@app.get("/api/descargar/alumnos")
def api_descargar_alumnos():
    ruta_archivo = os.path.join(os.path.dirname(__file__), "alumnos_sistema.xlsx")
    if os.path.exists(ruta_archivo):
        return FileResponse(ruta_archivo, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename="alumnos_sistema.xlsx")
    else:
        raise HTTPException(status_code=404, detail="Archivo no encontrado. Primero genera el Excel con POST /api/exportar/alumnos")

@app.get("/api/exportar/excel/filtrado")
def exportar_excel_filtrado(
    curso: Optional[str] = None,
    racion: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None
):
    """Genera y descarga un xlsx con los registros filtrados por curso, ración y rango de fechas."""
    import openpyxl
    from datetime import date
    from infra.DB.modelos import Registro
    from fastapi.responses import StreamingResponse
    import io

    db = SessionLocal()
    try:
        q = db.query(Registro).order_by(Registro.fecha.desc(), Registro.hora.desc())
        if fecha_desde:
            q = q.filter(Registro.fecha >= fecha_desde)
        if fecha_hasta:
            q = q.filter(Registro.fecha <= fecha_hasta)
        if racion:
            q = q.filter(Registro.tipo_solicitud == racion)
        registros = q.all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registros"
        ws.append(["ID", "Estudiante", "RUT", "Curso", "Ración", "Fecha", "Hora", "Terminal", "Estado"])

        for r in registros:
            u = r.usuario
            c = u.curso if u else None
            curso_nombre = (" ".join(p for p in [str(c.numero) if c and c.numero else "", c.nivel if c else "", c.letra or "" if c else ""] if p).strip()) if c else "Sin curso"
            if curso and curso_nombre != curso:
                continue
            ws.append([
                r.id,
                u.nombre if u else "Desconocido",
                u.rut if u else "",
                curso_nombre,
                r.tipo_solicitud,
                str(r.fecha) if r.fecha else "",
                str(r.hora) if r.hora else "",
                r.totem.ubicacion if r.totem else str(r.totem_id),
                r.estado_registro,
            ])

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        nombre_archivo = f"registros_{date.today()}.xlsx"
        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"}
        )
    finally:
        db.close()

@app.post("/api/exportar/excel")
async def api_exportar_excel():
    loop = asyncio.get_running_loop()
    resultado = await loop.run_in_executor(None, registros_controller.exportar_excel)
    if resultado["estado"]:
        return {"estado": True, "mensaje": "Excel generado exitosamente", "ruta": resultado["ruta"]}
    else:
        raise HTTPException(status_code=500, detail=resultado["mensaje"])
    
@app.get("/api/descargar/excel")
def api_descargar_excel():
    ruta_archivo = os.path.join(os.path.dirname(__file__), "registros.xlsx")
    if os.path.exists(ruta_archivo):
        return FileResponse(ruta_archivo, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename="registros.xlsx")
    else:
        raise HTTPException(status_code=404, detail="Archivo no encontrado")

@app.get("/api/exportar/asistencia")
def exportar_asistencia(anio: int, mes: int, curso: Optional[str] = None):
    """Genera y descarga un xlsx de asistencia JUNAEB estilo grilla mensual."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse
    import io
    from calendar import monthrange
    from datetime import date
    from infra.DB.modelos import Registro

    db = SessionLocal()
    try:
        # ── Alumnos ──────────────────────────────────────────────────────────
        def _curso_str(u):
            c = u.curso
            if not c:
                return "Sin curso"
            return " ".join(p for p in [str(c.numero) if c.numero is not None else "", c.nivel, c.letra or ""] if p).strip()

        usuarios = db.query(Usuario).all()
        if curso:
            usuarios = [u for u in usuarios if _curso_str(u) == curso]
        usuarios = sorted(usuarios, key=_sort_key_curso_chileno)

        # ── Registros del mes ────────────────────────────────────────────────
        dias_en_mes = monthrange(anio, mes)[1]
        f_ini = date(anio, mes, 1)
        f_fin = date(anio, mes, dias_en_mes)

        registros = db.query(Registro).filter(
            Registro.fecha >= f_ini,
            Registro.fecha <= f_fin,
            Registro.estado_registro == "Aprobado"
        ).all()

        asistencia: dict[int, set] = {}
        for r in registros:
            asistencia.setdefault(r.usuario_id, set()).add(r.fecha.day)

        # ── Workbook ─────────────────────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Asistencia JUNAEB"

        # Estilos
        FILL_TITLE   = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        FILL_HEADER  = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
        FILL_WEEKEND = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")
        FILL_ALT     = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        FILL_WHITE   = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        FONT_TITLE   = Font(bold=True, size=12, color="FFFFFF")
        FONT_MONTH   = Font(bold=True, size=10, color="000000")
        FONT_HEADER  = Font(bold=True, size=8,  color="FFFFFF")
        FONT_DATA    = Font(size=8,  color="000000")
        FONT_X       = Font(bold=True, size=8, color="000000")
        FONT_F       = Font(size=8, color="808080")
        ALIGN_C      = Alignment(horizontal="center", vertical="center")
        ALIGN_L      = Alignment(horizontal="left",   vertical="center", wrap_text=False)
        thin         = Side(style="thin", color="808080")
        BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Columnas: Nº(1) RUT(2) NOMBRE(3) CURSO(4) day1…dayN
        N_FIXED = 4
        last_cl = get_column_letter(N_FIXED + dias_en_mes)

        # ── Fila 1: Título ───────────────────────────────────────────────────
        ws.merge_cells(f"A1:{last_cl}1")
        c = ws["A1"]
        c.value     = f"COLEGIO PIAMARTA - ASISTENCIA JUNAEB {anio}"
        c.font      = FONT_TITLE
        c.fill      = FILL_TITLE
        c.alignment = ALIGN_C

        # ── Fila 2: Mes ──────────────────────────────────────────────────────
        mes_nombre = date(anio, mes, 1).strftime("%B %Y").upper()
        ws.merge_cells(f"A2:{last_cl}2")
        c = ws["A2"]
        c.value     = mes_nombre
        c.font      = FONT_MONTH
        c.alignment = ALIGN_C

        # ── Fila 3: Cabeceras ────────────────────────────────────────────────
        HDR_ROW = 3
        for col, txt in [(1, "Nº"), (2, "RUT"), (3, "NOMBRE"), (4, "CURSO")]:
            c = ws.cell(row=HDR_ROW, column=col, value=txt)
            c.font = FONT_HEADER; c.fill = FILL_HEADER
            c.alignment = ALIGN_C; c.border = BORDER

        for d in range(1, dias_en_mes + 1):
            col = N_FIXED + d
            dow = date(anio, mes, d).weekday()
            c = ws.cell(row=HDR_ROW, column=col, value=str(d))
            c.font = FONT_HEADER
            c.fill = FILL_WEEKEND if dow >= 5 else FILL_HEADER
            c.alignment = ALIGN_C; c.border = BORDER

        # ── Filas de datos ───────────────────────────────────────────────────
        for idx, u in enumerate(usuarios, 1):
            row = HDR_ROW + idx
            fill_row = FILL_ALT if idx % 2 == 0 else FILL_WHITE

            partes = u.nombre.strip().split()
            primer_nombre  = partes[0] if partes else ""
            apellido_pat   = partes[1] if len(partes) > 1 else ""
            nombre_display = f"{primer_nombre} {apellido_pat}".strip()
            curso_str = _curso_str(u)

            for col, val, align in [
                (1, idx,             ALIGN_C),
                (2, u.rut or "",     ALIGN_C),
                (3, nombre_display,  ALIGN_L),
                (4, curso_str,       ALIGN_L),
            ]:
                c = ws.cell(row=row, column=col, value=val)
                c.font = FONT_DATA; c.fill = fill_row
                c.alignment = align; c.border = BORDER

            dias_asistidos = asistencia.get(u.id, set())

            for d in range(1, dias_en_mes + 1):
                col = N_FIXED + d
                dow = date(anio, mes, d).weekday()
                asistio = d in dias_asistidos

                if dow >= 5:
                    c = ws.cell(row=row, column=col, value="F")
                    c.fill = FILL_WEEKEND
                    c.font = FONT_F
                elif asistio:
                    c = ws.cell(row=row, column=col, value="X")
                    c.fill = fill_row
                    c.font = FONT_X
                else:
                    c = ws.cell(row=row, column=col, value="")
                    c.fill = fill_row
                    c.font = FONT_DATA
                c.alignment = ALIGN_C; c.border = BORDER

        # ── Anchos de columna ────────────────────────────────────────────────
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 13
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 12
        for d in range(1, dias_en_mes + 1):
            ws.column_dimensions[get_column_letter(N_FIXED + d)].width = 3.2

        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 16
        ws.row_dimensions[HDR_ROW].height = 28

        # ── Stream ───────────────────────────────────────────────────────────
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        nombre_archivo = f"asistencia_junaeb_{anio}_{str(mes).zfill(2)}.xlsx"
        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"}
        )
    finally:
        db.close()


@app.post("/api/huella/vincular")
def api_vincular_huella(datos: VincularHuella):
    """Vincula una huella ya capturada a un usuario existente. Guarda en SQLite y en hardware."""
    from infra.DB.modelos import Huella
    db = SessionLocal()
    try:
        usuario = db.query(Usuario).filter(Usuario.id == datos.usuario_id).first()
        if not usuario:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")

        # Eliminar huella huérfana con los mismos bytes si existe
        db.query(Huella).filter(
            Huella.huella_blob == datos.huella_hex,
            Huella.usuario_id == None
        ).delete(synchronize_session=False)

        huella_existente = db.query(Huella).filter(Huella.usuario_id == datos.usuario_id).first()
        if huella_existente:
            huella_existente.huella_blob = datos.huella_hex
        else:
            db.add(Huella(huella_blob=datos.huella_hex, usuario_id=datos.usuario_id))
        db.commit()

        huella_bytes = bytes.fromhex(datos.huella_hex)
        hardware_service.guardar_en_bd(datos.usuario_id, huella_bytes)

        return {"estado": True, "mensaje": "Huella vinculada correctamente"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()

@app.post("/api/usuarios/importar")
async def api_importar_usuarios(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Archivo no válido. Se requiere un archivo Excel.")
    
    contenido = await file.read()
    loop = asyncio.get_running_loop()
    resultado = await loop.run_in_executor(None, lambda: registros_controller.importar_usuarios_desde_excel(contenido))
    
    if resultado["estado"]:
        return {"estado": True, "mensaje": resultado["mensaje"]}
    else:
        raise HTTPException(status_code=500, detail=resultado["mensaje"])

# ================================
# AUTH
# ================================

class LoginPayload(BaseModel):
    rut: str
    password: str

class CambiarCredenciales(BaseModel):
    rut_actual: str
    nuevo_rut: str
    nueva_password: str

@app.post("/api/auth/login")
def api_auth_login(datos: LoginPayload):
    db = SessionLocal()
    try:
        config = db.query(AdminConfig).first()
        if not config:
            raise HTTPException(status_code=500, detail="Configuración de admin no encontrada")
        hash_ingresado = _hash_password(datos.password)
        if datos.rut == config.rut and hash_ingresado == config.password_hash:
            return {"success": True}
        return {"success": False, "error": "Usuario o contraseña incorrectos"}
    finally:
        db.close()

@app.put("/api/auth/credenciales")
def api_cambiar_credenciales(datos: CambiarCredenciales):
    db = SessionLocal()
    try:
        config = db.query(AdminConfig).first()
        if not config:
            raise HTTPException(status_code=500, detail="Configuración de admin no encontrada")
        if datos.rut_actual.strip() != config.rut:
            raise HTTPException(status_code=401, detail="RUT incorrecto")
        if not datos.nuevo_rut.strip():
            raise HTTPException(status_code=400, detail="El nuevo usuario no puede estar vacío")
        if len(datos.nueva_password) < 6:
            raise HTTPException(status_code=400, detail="La nueva contraseña debe tener al menos 6 caracteres")
        config.rut = datos.nuevo_rut.strip()
        config.password_hash = _hash_password(datos.nueva_password)
        db.commit()
        return {"success": True, "mensaje": "Credenciales actualizadas correctamente"}
    finally:
        db.close()

app.mount("/", StaticFiles(directory=static_dir, html=True), name="frontend")
if __name__ == "__main__":
    uvicorn.run(app, host="127.0.0.1", port=8080)