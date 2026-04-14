import io
import os
import openpyxl
from datetime import date
from sqlalchemy.exc import IntegrityError as SAIntegrityError
from core.Domain.Repository.RegistroRepository import (
    crear_registro_evento,
    generar_ticket_comida,
)
from core.Domain.Repository.UserRepository import SessionLocal
from core.Services.ticketServices.horarioService import obtener_tipo_racion, descripcion_horario
from infra.DB.modelos import Usuario, Curso, Ticket


class RegistrosController:
    def __init__(self, totem_id: int):
        self.totem_id = totem_id

    def procesarAsistencia(self, usuario_id: int) -> dict:
        tipo_solicitud = obtener_tipo_racion()
        if tipo_solicitud is None:
            return {"estado": "Rechazado", "mensaje": f"Fuera de horario. {descripcion_horario()}"}

        db = SessionLocal()
        try:
            usuario = db.query(Usuario).filter(Usuario.id == usuario_id).first()

            if not usuario:
                return {"estado": "Rechazado", "mensaje": "Usuario no encontrado en el sistema"}

            if not usuario.es_pae:
                return {"estado": "Rechazado", "mensaje": "El alumno no pertenece al plan PAE"}

            ya_tiene = db.query(Ticket).filter(
                Ticket.usuario_id == usuario_id,
                Ticket.tipo_ticket == tipo_solicitud,
                Ticket.fecha_emision == date.today()
            ).first()
            if ya_tiene:
                return {"estado": "Rechazado", "mensaje": f"Ya se emitió un ticket de {tipo_solicitud} hoy"}

            registro = crear_registro_evento(usuario_id, self.totem_id, "Aprobado", tipo_solicitud)
            if not registro:
                return {"estado": "Rechazado", "mensaje": "Error al guardar el registro"}

            generar_ticket_comida(registro.id, usuario_id, tipo_solicitud)

            curso = (" ".join(p for p in [str(usuario.curso.numero) if usuario.curso.numero is not None else "", usuario.curso.nivel, usuario.curso.letra or ""] if p).strip()) if usuario.curso else "Sin curso"
            alumno = {"nombre": usuario.nombre, "rut": usuario.rut, "curso": curso}

            return {
                "estado": "Aprobado",
                "mensaje": f"Ticket de {tipo_solicitud} emitido correctamente",
                "tipo_racion": tipo_solicitud,
                "alumno": alumno
            }
        finally:
            db.close()

    def exportar_excel(self) -> dict:
        db = SessionLocal()
        try:
            tickets = db.query(Ticket).order_by(Ticket.fecha_emision.desc()).all()

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Registros Colación"
            ws.append(["ID", "Nombre", "RUT", "Curso", "Tipo", "Fecha", "Hora"])

            for t in tickets:
                u = t.usuario
                curso = (" ".join(p for p in [str(u.curso.numero) if u.curso.numero is not None else "", u.curso.nivel, u.curso.letra or ""] if p).strip()) if u and u.curso else "Sin curso"
                ws.append([
                    t.id,
                    u.nombre if u else "Desconocido",
                    u.rut if u else "",
                    curso,
                    t.tipo_ticket,
                    str(t.fecha_emision),
                    str(t.hora_emision),
                ])

            ruta = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "registros.xlsx"))
            wb.save(ruta)
            return {"estado": True, "ruta": ruta}
        except Exception as e:
            return {"estado": False, "mensaje": str(e)}
        finally:
            db.close()

    def exportar_alumnos_excel(self) -> dict:
        db = SessionLocal()
        try:
            usuarios = db.query(Usuario).order_by(Usuario.id).all()

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Alumnos"
            ws.append(["RUT", "Nombre", "Curso", "Estado", "PAE"])

            for u in usuarios:
                curso = (" ".join(p for p in [str(u.curso.numero) if u.curso.numero is not None else "", u.curso.nivel, u.curso.letra or ""] if p).strip()) if u.curso else "Sin curso"
                estado = u.estado.nombre_estado if u.estado else "Sin estado"
                ws.append([
                    u.rut or "",
                    u.nombre,
                    curso,
                    estado,
                    "Sí" if u.es_pae else "No",
                ])

            ruta = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "alumnos_sistema.xlsx"))
            wb.save(ruta)
            return {"estado": True, "ruta": ruta}
        except Exception as e:
            return {"estado": False, "mensaje": str(e)}
        finally:
            db.close()

    def importar_usuarios_desde_excel(self, contenido: bytes) -> dict:
        """
        Importa alumnos desde el Excel de JUNAEB.
        Columnas esperadas: run, nombre, apellido paterno, apellido materno, curso, (nivel), es_pae
        """
        db = SessionLocal()
        try:
            wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

            # Normalizar alias de columnas (nombres → nombre)
            headers = ["nombre" if h == "nombres" else h for h in headers]

            # Validar columnas mínimas requeridas
            requeridas = {"nombre", "curso"}
            faltantes = requeridas - set(headers)
            if faltantes:
                return {"estado": False, "mensaje": f"Columnas requeridas no encontradas: {faltantes}"}

            idx_nombre   = headers.index("nombre")
            idx_curso    = headers.index("curso")
            idx_ap_pat   = headers.index("apellido paterno") if "apellido paterno" in headers else None
            idx_ap_mat   = headers.index("apellido materno") if "apellido materno" in headers else None
            idx_run      = headers.index("run")    if "run"    in headers else (headers.index("rut") if "rut" in headers else None)
            idx_pae      = headers.index("es_pae") if "es_pae" in headers else None

            insertados = 0
            actualizados = 0
            omitidos = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                # Construir nombre completo
                nombre_base = str(row[idx_nombre]).strip() if row[idx_nombre] else ""
                ap_pat = str(row[idx_ap_pat]).strip() if idx_ap_pat is not None and row[idx_ap_pat] else ""
                ap_mat = str(row[idx_ap_mat]).strip() if idx_ap_mat is not None and row[idx_ap_mat] else ""
                nombre = " ".join(filter(None, [nombre_base, ap_pat, ap_mat]))
                if not nombre or nombre == "None":
                    continue

                rut = str(row[idx_run]).strip() if idx_run is not None and row[idx_run] else None

                # es_pae acepta "si", "sí", True, 1
                pae_raw = row[idx_pae] if idx_pae is not None else None
                if isinstance(pae_raw, str):
                    es_pae = pae_raw.lower().strip() in ("si", "sí", "yes", "true", "1")
                else:
                    es_pae = bool(pae_raw) if pae_raw is not None else True

                # Resolver curso: el Excel JUNAEB solo trae el nivel sin letra (ej: "Kinder", "1 Basico")
                curso_txt = str(row[idx_curso]).strip() if row[idx_curso] else None
                curso_id = None
                if curso_txt:
                    partes = curso_txt.split()
                    if partes[0].isdigit():
                        curso_numero = int(partes[0])
                        curso_nivel = " ".join(partes[1:])
                    else:
                        curso_numero = None
                        curso_nivel = curso_txt
                    curso_obj = db.query(Curso).filter(
                        Curso.numero == curso_numero, Curso.nivel == curso_nivel, Curso.letra == ""
                    ).first()
                    if not curso_obj:
                        curso_obj = Curso(numero=curso_numero, nivel=curso_nivel, letra="")
                        db.add(curso_obj)
                        db.flush()
                    curso_id = curso_obj.id

                # Upsert: por RUT primero, luego por nombre+curso
                existente = None
                if rut:
                    existente = db.query(Usuario).filter(Usuario.rut == rut).first()
                if not existente and curso_id:
                    existente = db.query(Usuario).filter(
                        Usuario.nombre == nombre, Usuario.curso_id == curso_id
                    ).first()

                try:
                    sp = db.begin_nested()
                    if existente:
                        existente.nombre = nombre
                        existente.es_pae = es_pae
                        if rut:
                            existente.rut = rut
                        if curso_id:
                            existente.curso_id = curso_id
                        sp.commit()
                        actualizados += 1
                    else:
                        db.add(Usuario(nombre=nombre, rut=rut, curso_id=curso_id, estado_id=1, es_pae=es_pae))
                        db.flush()
                        sp.commit()
                        insertados += 1
                except SAIntegrityError:
                    sp.rollback()
                    omitidos += 1

            db.commit()
            return {
                "estado": True,
                "mensaje": f"{insertados} nuevos, {actualizados} actualizados, {omitidos} omitidos por duplicado",
                "insertados": insertados,
                "actualizados": actualizados,
                "omitidos": omitidos,
            }
        except Exception as e:
            db.rollback()
            return {"estado": False, "mensaje": f"Error al procesar el archivo: {str(e)}"}
        finally:
            db.close()
